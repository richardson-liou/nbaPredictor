from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import automationExample
from openpyxl import Workbook
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import pandas as pd

def getRows():
    url = "https://www.espn.com/nba/teams/comparison/_/team1/atl/team2/bos"
    path = "C:\\Users\\richa\\Downloads\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe"

    options = Options()
    options.add_argument("--headless")

    service = Service(executable_path=path)
    driver = webdriver.Chrome(service=service, options=options)

    driver.get(url)

    var = "//table[@class='tablehead']//tbody"
    containers = driver.find_elements(by="xpath", value=var)
    list = []
    if len(containers) > 1 :

        # print("There is more than one tbody we are getting to the second ")
        second_tbody = containers[2]
        rows = second_tbody.find_elements(by="xpath", value=".//tr")
        for row in rows:
            list.append(row.text)

    else:
        print("There is only one for some reason ")
    driver.quit()
    return list



teams = []
stats = []
players = []
values = []

wb = Workbook()

ws = wb.active
data = getRows()
statsData = getRows()[2:]
ws.title = data[0]

pattern =r"(\w+ \d{1,2})\s+([A-Za-z]+ \d{1,3},)\s+([A-Za-z]+ \d{1,3}(?: \(OT\))?)\s+(\w+ \d+ Pts)\s+(\w+ \d+ Pts)"
print(data)





final_data = []

headers = ["Date", "Team 1", "Team 2", "Team 1 Top Scorer", "Team 2 Top Scorer"]

ws.append(headers)

for i in data[2:-1]:
    row = []
    match = re.match(pattern, i)
    print(i)
    print("match:",match)
    if match:
        row.append(match.group(1))
        row.append( match.group(2))
        row.append( match.group(3))
        row.append( match.group(4))
        row.append( match.group(5))
    final_data.append(row)



for i in final_data:
    ws.append(i)
    print(i)

#
# for i in statsData:
#     parts = i.split()
#     stat = parts[0] + " " + parts[1] if "%" in parts[1] else parts[0]
#
#     player1 = parts[1] if "%" not in parts[1] else parts[2] + " " + parts[3]
#
#     value1 = parts[-4]
#     player2 = parts[-3] + " " + parts[-2]
#     value2 = parts[-1]
#
#     stats.append(stat)
#     players.append(player1)
#     values.append(value1)
#     teams.append("ATLANTA")
#
#     teams.append("BOSTON")
#     stats.append(stat)
#     players.append(player2)
#     values.append(value2)
#
# newData = {"stats":stats, "teams":teams, "values":values, "players":players}
# df= pd.DataFrame(newData)
#
#
# for i in df:
#     ws2.append(i)

wb.save("y.xlsx")