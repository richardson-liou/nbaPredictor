from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import automationExample
from openpyxl import Workbook
import re
import itertools
import pandas as pd


from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

def getRows(url, container):
    path = "C:\\Users\\richa\\Downloads\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe"


    options = Options()
    options.add_argument("--headless")

    service = Service(executable_path=path)
    driver = webdriver.Chrome(service=service, options=options)
    print("#####################")
    print(url)
    driver.get(url)

    var = "//table[@class='tablehead']//tbody"
    containers = driver.find_elements(by="xpath", value=var)
    list = []
    if len(containers) > 1 :

        # print("There is more than one tbody we are getting to the second ")
        second_tbody = containers[container]
        rows = second_tbody.find_elements(by="xpath", value=".//tr")
        for row in rows:
            list.append(row.text)
        driver.quit()
        return list

    driver.quit()
    return []

# print(getRows("https://www.espn.com/nba/teams/comparison/_/team1/atl/team2/bos"))

def getTeam():
    teamAbv = ['dal','den', 'det', 'gsw']
    # ,'cha','chi','chi','cle','dal','den','det','gsw','hou','ind','lac','lal',]
               # 'mem','mia','mil','min','nor','bkn','okc','orl','phi','pho','por','sac','tor','uth','was']

    teamPairs = list(itertools.combinations(teamAbv, 2))
    # return teamPairs
    urls = []
    for i in teamPairs:
        team1 = i[0]
        team2 = i[1]
        url = f'https://www.espn.com/nba/teams/comparison/_/year/2024/team1/{team1}/team2/{team2}'
        urls.append(url)
    return urls

teamsUrls = getTeam()


def writeDatatoExcel():
    wb = Workbook()
    ws = wb.active
    # ws.title = "Website Data"
    # Process this further
    headers = ["Date", "Team 1", "Team 2", "Team 1 Top Scorer", "Team 2 Top Scorer"]

    for i in teamsUrls:

        teams = []
        stats = []
        players = []
        values = []

        team_match_up_data =  getRows(i,1)

        print(f"Writing stat_leader_data for team url{i}")
        pattern = r"(\w+ \d{1,2})\s+([A-Za-z\s]+ \d{1,3}),\s+([A-Za-z\s]+ \d{1,3}(?: \(OT\))?)\s+([\w\s\.-]+ \d+ Pts)\s+([\w\s\.-]+ \d+ Pts)(?:\s+([\w\s\.-]+ \d+ Pts))?"

        if len(team_match_up_data) > 1:
            final_data = []

            for j in team_match_up_data[2:-1]:
                final_data.append(j)

            match = re.match(pattern, final_data[0])

            team1 = match.group(2).strip().split()[0]
            team2 = match.group(3).strip().split()[0]
            print(f"Team 1: {team1}, Team 2: {team2}")


        sheetName = f"Matchups {team1}-{team2}"
        ws = wb.create_sheet(title=sheetName)

        processed_data = []

        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        current_row = 2

        for p in final_data:
            print("Data P:",p)
            match = re.match(pattern, p)
            print("Match:",match)


            if match:
                date = match.group(1)
                team1_score = match.group(2)
                team2_score = match.group(3)
                top_scorer1 = match.group(4)
                top_scorer2 = match.group(5)

                # Write data to row 2 starting at column A
                data = [date, team1_score, team2_score, top_scorer1, top_scorer2]
                for col_num, value in enumerate(data, start=1):
                    ws.cell(row=current_row, column=col_num, value=value)
            current_row += 1
            # Save the workbook



        print("STAT LEADERS: ")


        stat_leader_data = getRows(i,2)[2:]
        print("STATS LEADER ")
        print(stat_leader_data)

        for i in stat_leader_data:
            parts = i.split()
            stat = parts[0] + " " + parts[1] if "%" in parts[1] else parts[0]

            player1 = parts[1] if "%" not in parts[1] else parts[2] + " " + parts[3]

            value1 = parts[-4]
            player2 = parts[-3] + " " + parts[-2]
            value2 = parts[-1]

            stats.append(stat)
            players.append(player1)
            values.append(value1)
            teams.append(team1)

            teams.append(team2)
            stats.append(stat)
            players.append(player2)
            values.append(value2)

            newData = {"stats": stats, "teams": teams, "values": values, "players": players}
            # df = pd.DataFrame(newData)



        sheetName = f"Stat Leaders_{team1}_{team2}"
        new_ws = wb.create_sheet(title = sheetName)

        headers2 = ["stat", "team", "value", "player"]

        new_ws.append(headers2)
        print("FINAL DATA")
        print(final_data)



        column_positions = {"stats": 1, "teams":2, "values":3 , "players":4 }

        start_row = 2
        start_col = 0



        
        for key, stat_leader_data in newData.items():
            c = column_positions.get(key,1)
            for i, value in enumerate(stat_leader_data):
                new_ws.cell(row = start_row + i, column = start_col + c, value = value )


    wb.save("nba_all_data.xlsx")

writeDatatoExcel()
